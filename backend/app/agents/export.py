"""
Export Agent - экспорт данных в различные форматы.
"""
import logging
import csv
import io
from typing import List, Dict, Any, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExportAgent:
    """Agent для экспорта данных."""
    
    def __init__(self):
        """Initialize export agent."""
        pass
    
    def export_to_csv(
        self,
        posts: List[Dict[str, Any]],
        columns: Optional[List[str]] = None
    ) -> bytes:
        """
        Export posts to CSV format.
        
        Args:
            posts: List of post dictionaries
            columns: List of column names to export (None = all)
            
        Returns:
            CSV file as bytes
        """
        if not posts:
            return b""
        
        # Default columns if not specified
        if columns is None:
            columns = [
                "id", "post_id", "channel", "date", "text", "author",
                "views", "likes", "engagement_rate", "content_type",
                "hashtags", "mentions", "links"
            ]
        
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(columns)
        
        # Write data rows
        for post in posts:
            row = []
            for col in columns:
                value = post.get(col)
                
                # Handle special cases
                if col == "hashtags" or col == "mentions" or col == "links":
                    if isinstance(value, list):
                        value = ", ".join(value) if value else ""
                    else:
                        value = ""
                elif col == "date":
                    if isinstance(value, datetime):
                        value = value.isoformat()
                    elif isinstance(value, str):
                        value = value  # Already formatted
                    else:
                        value = ""
                elif value is None:
                    value = ""
                else:
                    value = str(value)
                
                row.append(value)
            
            writer.writerow(row)
        
        # Convert to bytes
        csv_bytes = output.getvalue().encode('utf-8-sig')  # BOM for Excel compatibility
        output.close()
        
        return csv_bytes
    
    def export_to_excel(
        self,
        posts: List[Dict[str, Any]],
        columns: Optional[List[str]] = None
    ) -> bytes:
        """
        Export posts to Excel format.
        
        Args:
            posts: List of post dictionaries
            columns: List of column names to export (None = all)
            
        Returns:
            Excel file as bytes
        """
        if not posts:
            return b""
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Posts"
        
        # Default columns if not specified
        if columns is None:
            columns = [
                "id", "post_id", "channel", "date", "text", "author",
                "views", "likes", "engagement_rate", "content_type",
                "hashtags", "mentions", "links"
            ]
        
        # Write header
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=idx, value=col)
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data rows
        for row_idx, post in enumerate(posts, start=2):
            for col_idx, col in enumerate(columns, start=1):
                value = post.get(col)
                
                # Handle special cases
                if col == "hashtags" or col == "mentions" or col == "links":
                    if isinstance(value, list):
                        value = ", ".join(value) if value else ""
                    else:
                        value = ""
                elif col == "date":
                    if isinstance(value, datetime):
                        value = value
                    elif isinstance(value, str):
                        try:
                            value = datetime.fromisoformat(value.replace('Z', '+00:00'))
                        except:
                            value = value
                    else:
                        value = ""
                elif value is None:
                    value = ""
                
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for idx, col in enumerate(columns, start=1):
            max_length = len(str(col))
            for row in ws.iter_rows(min_row=2, max_row=len(posts) + 1, min_col=idx, max_col=idx):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            
            # Limit max width
            max_length = min(max_length, 50)
            ws.column_dimensions[get_column_letter(idx)].width = max_length + 2
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        excel_bytes = output.getvalue()
        output.close()
        
        return excel_bytes
    
    def export_to_google_sheets_format(
        self,
        posts: List[Dict[str, Any]],
        columns: Optional[List[str]] = None
    ) -> List[List[Any]]:
        """
        Format posts for Google Sheets API.
        
        Args:
            posts: List of post dictionaries
            columns: List of column names to export (None = all)
            
        Returns:
            List of lists (rows) ready for Google Sheets API
        """
        if not posts:
            return []
        
        # Default columns if not specified
        if columns is None:
            columns = [
                "id", "post_id", "channel", "date", "text", "author",
                "views", "likes", "engagement_rate", "content_type",
                "hashtags", "mentions", "links"
            ]
        
        # Prepare rows
        rows = []
        
        # Header row
        rows.append(columns)
        
        # Data rows
        for post in posts:
            row = []
            for col in columns:
                value = post.get(col)
                
                # Handle special cases
                if col == "hashtags" or col == "mentions" or col == "links":
                    if isinstance(value, list):
                        value = ", ".join(value) if value else ""
                    else:
                        value = ""
                elif col == "date":
                    if isinstance(value, datetime):
                        value = value.isoformat()
                    elif isinstance(value, str):
                        value = value
                    else:
                        value = ""
                elif value is None:
                    value = ""
                else:
                    value = str(value)
                
                row.append(value)
            
            rows.append(row)
        
        return rows
    
    def prepare_posts_for_export(
        self,
        posts: List[Any],
        include_channel_info: bool = True
    ) -> List[Dict[str, Any]]:
        """
        Prepare post objects for export.
        
        Args:
            posts: List of Post model objects
            include_channel_info: Include channel name and avatar
            
        Returns:
            List of dictionaries ready for export
        """
        export_data = []
        
        for post in posts:
            data = {
                "id": post.id,
                "post_id": post.post_id,
                "date": post.date,
                "text": post.text,
                "author": post.author,
                "views": post.views,
                "likes": post.likes,
                "engagement_rate": post.engagement_rate,
                "content_type": post.content_type,
                "hashtags": post.hashtags or [],
                "mentions": post.mentions or [],
                "links": post.links or [],
            }
            
            if include_channel_info and post.channel:
                data["channel"] = post.channel.channel_name
                data["channel_avatar"] = post.channel.channel_avatar_url
            else:
                data["channel"] = None
                data["channel_avatar"] = None
            
            export_data.append(data)
        
        return export_data
    
    def export_streaming(
        self,
        posts_generator,
        format: str = "csv",
        columns: Optional[List[str]] = None
    ):
        """
        Stream export for large datasets.
        
        Args:
            posts_generator: Generator yielding post dictionaries
            format: Export format ("csv" or "excel")
            columns: List of column names to export
            
        Yields:
            Chunks of exported data
        """
        if format == "csv":
            # CSV header
            if columns is None:
                columns = [
                    "id", "post_id", "channel", "date", "text", "author",
                    "views", "likes", "engagement_rate", "content_type",
                    "hashtags", "mentions", "links"
                ]
            
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(columns)
            yield output.getvalue().encode('utf-8-sig')
            output.close()
            
            # Stream data rows
            for post in posts_generator:
                output = io.StringIO()
                writer = csv.writer(output)
                row = []
                for col in columns:
                    value = post.get(col)
                    if isinstance(value, list):
                        value = ", ".join(value) if value else ""
                    elif value is None:
                        value = ""
                    else:
                        value = str(value)
                    row.append(value)
                writer.writerow(row)
                yield output.getvalue().encode('utf-8')
                output.close()
        else:
            # For Excel, we need to collect all data first
            # Or use a library that supports streaming Excel
            raise NotImplementedError("Streaming Excel export not implemented")

